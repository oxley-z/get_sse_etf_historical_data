"""
上交所宽基 ETF 规模监控（断点续传 + 断网保存 + 增量更新）
依赖安装：pip install requests plotly openpyxl

运行方式：python sse_etf_dashboard.py
脚本启动时自动判断运行模式：

  【历史下载模式】无 checkpoint 或历史尚未抓完
    从今天/断点起向过去方向逐日抓至 2020-01-01
    每成功一天立即写断点；Ctrl+C 或断网自动保存

  【增量更新模式】checkpoint 存在且 note 含"完成"
    找到已有数据的最新日期，从次日起正向抓到今天
    新数据追加合并，重新生成 HTML + Excel

输出：sse_final_dashboard.html  交互图表
      sse_etf_data.xlsx         历史数据（透视表 + 明细表）
      sse_checkpoint.json       进度文件（永久保留）
"""

import requests
import json
import time
import os
import re
import webbrowser
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 1. 配置 ────────────────────────────────────────────────────────────────────

ETF_MAP = {
    '510300': '华泰柏瑞沪深300ETF',
    '510310': '易方达沪深300ETF',
    '510330': '华夏沪深300ETF',
    '510050': '华夏上证50ETF',
    '510500': '南方中证500ETF',
    '512100': '南方中证1000ETF',
    '510180': '华安上证180ETF',
    '560010': '广发中证1000ETF',
    '588080': '易方达上证科创板50ETF',
}

TARGET_DAYS      = 1500
CUTOFF_DATE      = datetime(2020, 1, 1)
OUTPUT_HTML      = 'sse_final_dashboard.html'
OUTPUT_EXCEL     = 'sse_etf_data.xlsx'
CHECKPOINT       = 'sse_checkpoint.json'
MAX_NET_FAILURES = 5
NET_RETRY_WAIT   = 3

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'Referer':    'https://www.sse.com.cn/',
    'Accept':     '*/*',
}


# 上证指数配置
SHINDEX_CHECKPOINT = 'shindex_checkpoint.json'   # 上证指数独立 checkpoint
SHINDEX_SOHU_URL   = (
    'http://q.stock.sohu.com/hisHq'
    '?code=zs_000001'
    '&start=20200101'
    '&end=99991231'
    '&stat=1&order=D&period=d'
    '&callback=historySearchHandler&rt=jsonp'
)
SHINDEX_HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'Referer':    'http://q.stock.sohu.com/',
    'Accept':     '*/*',
}

# ── 2. 断点工具 ────────────────────────────────────────────────────────────────

def load_checkpoint():
    if not os.path.exists(CHECKPOINT):
        return [], None
    try:
        with open(CHECKPOINT, 'r', encoding='utf-8') as f:
            data = json.load(f)
        results   = data.get('results', [])
        last_date = data.get('last_date', None)
        note      = data.get('note', '')
        print(f'📂 断点文件：{len(results)} 个交易日 | 最早 {last_date} | 备注: {note}')
        return results, last_date
    except Exception as e:
        print(f'  [警告] 断点文件读取失败，重新开始: {e}')
        return [], None


def save_checkpoint(results, note=''):
    """原子写入断点文件，永久保留。"""
    if not results:
        return
    dates     = [r['date'] for r in results]
    last_date = min(dates)   # 最早（历史最远）
    first_date = max(dates)  # 最新（最近今天）
    tmp = CHECKPOINT + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump({
            'last_date':  last_date,
            'first_date': first_date,
            'count':      len(results),
            'saved_at':   datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'note':       note,
            'results':    results,
        }, f, ensure_ascii=False)
    os.replace(tmp, CHECKPOINT)


def today_str():
    """返回当天日期字符串 YYYY-MM-DD。"""
    return datetime.today().strftime('%Y-%m-%d')


def get_existing_dates():
    """返回 checkpoint 中所有已有日期的集合，文件不存在时返回空集合。"""
    if not os.path.exists(CHECKPOINT):
        return set()
    try:
        with open(CHECKPOINT, 'r', encoding='utf-8') as f:
            results = json.load(f).get('results', [])
        return {r['date'] for r in results}
    except Exception:
        return set()


def read_latest_date():
    """读取 checkpoint 中最新（最大）的日期。"""
    try:
        with open(CHECKPOINT, 'r', encoding='utf-8') as f:
            return json.load(f).get('first_date', None)
    except Exception:
        return None


def read_all_results():
    try:
        with open(CHECKPOINT, 'r', encoding='utf-8') as f:
            return json.load(f).get('results', [])
    except Exception:
        return []

# ── 3. 网络请求 ────────────────────────────────────────────────────────────────

def is_network_error(e):
    msg = str(e).lower()
    return any(k in msg for k in (
        'connectionerror', 'timeout', 'max retries', 'connection reset',
        'remotedisconnected', 'network', 'proxy', 'nodename nor servname',
        'name or service not known', 'failed to establish',
    ))


def fetch_day(date_str):
    """
    返回 (items,'ok') | (None,'nodata') | (None,'neterr') | (None,'apierr')
    """
    ts  = int(time.time() * 1000)
    url = (
        'https://query.sse.com.cn/commonQuery.do'
        f'?isPagination=true&pageHelp.pageSize=1000'
        f'&sqlId=COMMON_SSE_ZQPZ_ETFZL_XXPL_ETFGM_SEARCH_L'
        f'&STAT_DATE={date_str}&_{ts}'
    )
    try:
        resp  = requests.get(url, headers=HEADERS, timeout=10,
                             proxies={'http': None, 'https': None})
        resp.raise_for_status()
        items = resp.json().get('pageHelp', {}).get('data', [])
        return (items, 'ok') if items else (None, 'nodata')
    except Exception as e:
        return (None, 'neterr') if is_network_error(e) else (None, 'apierr')


# ── 4. 通用抓取循环（供两种模式复用） ─────────────────────────────────────────

def fetch_dates(date_list, existing_dates, mode_label, is_recent=False):
    """
    对 date_list 中尚未在 existing_dates 里的日期逐一抓取。
    is_recent=True 时，nodata 提示改为"数据未发布/非交易日"，避免误判。
    返回 (new_records列表, stop_reason)
    stop_reason: 'completed' | 'network' | 'interrupt'
    """
    new_records    = []
    net_fail_count = 0
    stop_reason    = 'completed'

    try:
        for date_str in date_list:
            if date_str in existing_dates:
                continue

            print(f'同步 {date_str} [{mode_label}，已新增 {len(new_records)} 天]', end='  ')
            items, status = fetch_day(date_str)

            if status == 'ok':
                new_records.append({'date': date_str, 'items': items})
                net_fail_count = 0
                print('✓')

            elif status == 'nodata':
                net_fail_count = 0
                # 增量模式下接口可能存在数据延迟，不能简单判定为非交易日
                if is_recent:
                    print('— (数据未发布或非交易日)')
                else:
                    print('— (非交易日)')

            elif status == 'neterr':
                net_fail_count += 1
                print(f'✗ 网络异常 ({net_fail_count}/{MAX_NET_FAILURES})')
                if net_fail_count >= MAX_NET_FAILURES:
                    stop_reason = 'network'
                    break
                time.sleep(NET_RETRY_WAIT)
                continue   # 不移动，重试同一天

            else:
                print('— (API错误，跳过)')

            time.sleep(0.4)

    except KeyboardInterrupt:
        stop_reason = 'interrupt'
        print('\n\n⏸  Ctrl+C 捕获')

    return new_records, stop_reason


# ── 5. 历史下载模式 ────────────────────────────────────────────────────────────

def collect_history():
    """
    从今天/断点起向过去方向逐日抓取，直到 CUTOFF_DATE 或 TARGET_DAYS。
    返回 (results, completed布尔)
    """
    results, last_date = load_checkpoint()
    existing_dates     = {r['date'] for r in results}

    if last_date:
        start = datetime.strptime(last_date, '%Y-%m-%d') - timedelta(days=1)
        print(f'▶  历史下载（续传），从 {start.strftime("%Y-%m-%d")} 继续向前')
    else:
        start = datetime.today()
        print(f'▶  历史下载（首次），从今天起向前追溯')

    print(f'   目标 {TARGET_DAYS} 个交易日 | 截止 {CUTOFF_DATE.strftime("%Y-%m-%d")}')
    print('   Ctrl+C 可随时中断并保存进度')
    print('=' * 60)

    # 生成待抓日期列表（倒序：从 start 到 CUTOFF_DATE）
    date_list = []
    d = start
    while d >= CUTOFF_DATE and (len(existing_dates) + len(date_list)) < TARGET_DAYS:
        date_list.append(d.strftime('%Y-%m-%d'))
        d -= timedelta(days=1)

    new_records, stop_reason = fetch_dates(date_list, existing_dates, '历史')

    # 合并
    for r in new_records:
        results.append(r)
        existing_dates.add(r['date'])
        save_checkpoint(results, note='运行中')   # 实时持久化

    completed = (stop_reason == 'completed')

    if stop_reason == 'interrupt':
        save_checkpoint(results, note='Ctrl+C 中断')
        print(f'✅ 已保存 {len(results)} 个交易日 → {CHECKPOINT}，下次运行自动续传')
    elif stop_reason == 'network':
        save_checkpoint(results, note='断网自动保存')
        print(f'🔌 断网，已保存 {len(results)} 个交易日 → {CHECKPOINT}，恢复网络后重新运行')
    else:
        save_checkpoint(results, note='全量采集完成')
        print(f'✅ 历史数据采集完成，共 {len(results)} 个交易日，断点文件永久保留')

    return results, completed


# ── 6. 增量更新模式 ────────────────────────────────────────────────────────────

def incremental_update():
    """
    增量更新：从今天起往前逐日检查，遇到已在 checkpoint 中的日期即停止。
    将新抓到的数据追加合并后保存。
    返回 (merged_results列表, 新增天数)
    """
    existing_results = read_all_results()
    existing_dates   = {r['date'] for r in existing_results}

    # 从今天起向前生成候选日期，遇到已有日期就停（最多往前看 30 个自然日防死循环）
    candidates = []
    d = datetime.today()
    for _ in range(30):
        ds = d.strftime('%Y-%m-%d')
        if ds in existing_dates:
            break          # 碰到已有数据，停止往前看
        candidates.append(ds)
        d -= timedelta(days=1)

    if not candidates:
        print(f'   最新数据已是今天，无需更新。')
        return existing_results, 0

    print(f'   待检查日期: {candidates[-1]} ~ {candidates[0]}（{len(candidates)} 个自然日）')
    print('   Ctrl+C 可随时中断，已抓到的新数据会保存')
    print('=' * 60)

    # 正序抓取（从旧到新），便于连续失败时已有较早数据
    date_list = list(reversed(candidates))

    new_records, stop_reason = fetch_dates(date_list, existing_dates, '增量', is_recent=True)

    # 合并去重，保持倒序（与历史下载一致）
    merged_map = {r['date']: r for r in existing_results}
    for r in new_records:
        merged_map[r['date']] = r
    merged_results = sorted(merged_map.values(), key=lambda x: x['date'], reverse=True)

    new_count = len(new_records)
    if new_count > 0:
        if stop_reason == 'completed':
            note = f'增量更新完成，新增 {new_count} 天'
        else:
            note = f'增量更新中断（{stop_reason}），已新增 {new_count} 天'
        save_checkpoint(merged_results, note=note)
        print(f'\n✅ 新增 {new_count} 个交易日，已合并保存 → {CHECKPOINT}')
    else:
        if stop_reason == 'completed':
            print(f'\n✅ 检查完毕，期间暂无新交易日数据（可能是节假日或数据尚未发布）。')
        else:
            print(f'\n⚠  更新中断（{stop_reason}），未获取到新数据。')

    return merged_results, new_count


def collect_history():
    """
    从今天/断点起向过去方向逐日抓取，直到 CUTOFF_DATE 或 TARGET_DAYS。
    返回 (results, completed布尔)
    """
    results, last_date = load_checkpoint()
    existing_dates     = {r['date'] for r in results}

    if last_date:
        start = datetime.strptime(last_date, '%Y-%m-%d') - timedelta(days=1)
        print(f'▶  历史下载（续传），从 {start.strftime("%Y-%m-%d")} 继续向前')
    else:
        start = datetime.today()
        print(f'▶  历史下载（首次），从今天起向前追溯')

    print(f'   目标 {TARGET_DAYS} 个交易日 | 截止 {CUTOFF_DATE.strftime("%Y-%m-%d")}')
    print('   Ctrl+C 可随时中断并保存进度')
    print('=' * 60)

    # 生成待抓日期列表（倒序：从 start 到 CUTOFF_DATE）
    date_list = []
    d = start
    while d >= CUTOFF_DATE and (len(existing_dates) + len(date_list)) < TARGET_DAYS:
        date_list.append(d.strftime('%Y-%m-%d'))
        d -= timedelta(days=1)

    new_records, stop_reason = fetch_dates(date_list, existing_dates, '历史')

    # 合并
    for r in new_records:
        results.append(r)
        existing_dates.add(r['date'])
        save_checkpoint(results, note='运行中')   # 实时持久化

    completed = (stop_reason == 'completed')

    if stop_reason == 'interrupt':
        save_checkpoint(results, note='Ctrl+C 中断')
        print(f'✅ 已保存 {len(results)} 个交易日 → {CHECKPOINT}，下次运行自动续传')
    elif stop_reason == 'network':
        save_checkpoint(results, note='断网自动保存')
        print(f'🔌 断网，已保存 {len(results)} 个交易日 → {CHECKPOINT}，恢复网络后重新运行')
    else:
        save_checkpoint(results, note='全量采集完成')
        print(f'✅ 历史数据采集完成，共 {len(results)} 个交易日，断点文件永久保留')

    return results, completed


# ── 6. 增量更新模式 ────────────────────────────────────────────────────────────

def incremental_update():
    """
    找到已有数据最新日期，从次日起正向抓到今天，追加合并后保存。
    返回 (merged_results, new_count)
    """
    latest_str = read_latest_date()
    if not latest_str:
        print('❌ 无法读取已有最新日期，请检查断点文件。')
        return [], 0

    latest = datetime.strptime(latest_str, '%Y-%m-%d')
    today  = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)

    if latest >= today:
        print(f'✅ 数据已是最新（{latest_str}），无需更新。')
        return read_all_results(), 0

    # 生成待检查日期（正向：latest+1 → today）
    date_list, d = [], latest + timedelta(days=1)
    while d <= today:
        date_list.append(d.strftime('%Y-%m-%d'))
        d += timedelta(days=1)

    print(f'▶  增量更新模式')
    print(f'   已有数据最新日期 : {latest_str}')
    print(f'   待检查范围       : {date_list[0]} ~ {date_list[-1]}（{len(date_list)} 个自然日）')
    print('   Ctrl+C 可随时中断，已抓到的新数据会保存')
    print('=' * 60)

    existing_results = read_all_results()
    existing_dates   = {r['date'] for r in existing_results}

    new_records, stop_reason = fetch_dates(date_list, existing_dates, '增量', is_recent=True)

    # 合并去重，保持倒序
    merged_map = {r['date']: r for r in existing_results}
    for r in new_records:
        merged_map[r['date']] = r
    merged_results = sorted(merged_map.values(), key=lambda x: x['date'], reverse=True)

    new_count = len(new_records)
    if new_count > 0:
        if stop_reason == 'completed':
            note = f'增量更新完成，新增 {new_count} 天'
        else:
            note = f'增量更新中断（{stop_reason}），已新增 {new_count} 天'
        save_checkpoint(merged_results, note=note)
        print(f'\n✅ 新增 {new_count} 个交易日，已合并保存 → {CHECKPOINT}')
    else:
        if stop_reason == 'completed':
            print(f'\n✅ 检查完毕，期间暂无新交易日数据。')
        else:
            print(f'\n⚠  更新中断（{stop_reason}），未获取到新数据。')

    return merged_results, new_count


# ── 7. 数据整理 ────────────────────────────────────────────────────────────────

def parse_val(v):
    if v is None:
        return None
    try:
        return float(str(v).replace(',', ''))
    except ValueError:
        return None


def sniff_keys(sample):
    keys     = list(sample.keys())
    code_key = next(
        (k for k in keys if str(sample[k]).strip()[:2] in ('51', '56', '58')),
        'SEC_CODE'
    )
    num_keys = [k for k in keys if k != code_key
                and 'DATE' not in k.upper()
                and parse_val(sample[k]) is not None]
    val_key = (
        next((k for k in num_keys if any(kw in k.upper()
              for kw in ('VOL', 'FE', 'SHARE', '份额', '总量'))), None)
        or next((k for k in num_keys if any(kw in k.upper()
              for kw in ('VAL', 'SZ', '市值'))), None)
        or (num_keys[0] if num_keys else None)
    )
    return code_key, val_key


def build_plot_data(results, index_prices=None):
    if not results:
        return []
    code_key, val_key = sniff_keys(results[0]['items'][0])
    print(f'字段嗅探 → 代码: {code_key}  |  数值: {val_key}')
    plot_data = []

    # ── ETF traces（左轴 yaxis） ───────────────────────────────────────────────
    for code, name in ETF_MAP.items():
        dates, values = [], []
        for day in results:
            item = next(
                (i for i in day['items'] if str(i.get(code_key, '')).strip() == code),
                None
            )
            dates.append(day['date'])
            values.append(parse_val(item[val_key]) if item else None)
        pairs = sorted(zip(dates, values), key=lambda x: x[0])
        plot_data.append({
            'x': [p[0] for p in pairs], 'y': [p[1] for p in pairs],
            'name': f'{name}({code})', 'mode': 'lines+markers',
            'line': {'width': 2.5}, 'marker': {'size': 6}, 'connectgaps': False,
            'yaxis': 'y',   # 左轴
        })

    # ── 上证指数 trace（右轴 yaxis2） ─────────────────────────────────────────
    if index_prices:
        # 只保留与 ETF 数据日期范围重叠的部分，避免图表过宽
        all_etf_dates = {d for t in plot_data for d in t['x']}
        if all_etf_dates:
            min_date = min(all_etf_dates)
            max_date = max(all_etf_dates)
            idx_pairs = sorted(
                [(d, v) for d, v in index_prices.items() if min_date <= d <= max_date],
                key=lambda x: x[0]
            )
        else:
            idx_pairs = sorted(index_prices.items())

        if idx_pairs:
            plot_data.append({
                'x':    [p[0] for p in idx_pairs],
                'y':    [p[1] for p in idx_pairs],
                'name': '上证指数(000001)',
                'mode': 'lines',
                'line': {'width': 2, 'color': '#e74c3c', 'dash': 'solid'},
                'yaxis': 'y2',  # 右轴
                'connectgaps': False,
                'opacity': 0.85,
            })
    return plot_data




# ── 上证指数数据抓取 ────────────────────────────────────────────────────────────

def fetch_shindex():
    """
    一次性从搜狐接口拉取上证指数 2020-01-01 至今的全量历史数据。
    返回 dict: {'YYYY-MM-DD': close_price, ...}  或 {} 若失败
    """
    try:
        resp = requests.get(SHINDEX_SOHU_URL, headers=SHINDEX_HEADERS,
                            timeout=20, proxies={'http': None, 'https': None})
        resp.raise_for_status()
        # JSONP 解包：historySearchHandler([{...}])
        text = resp.text.strip()
        m    = re.search(r'historySearchHandler\((.+)\)\s*;?\s*$', text, re.DOTALL)
        if not m:
            print('  [上证指数] JSONP 解析失败，原始响应前100字符:', text[:100])
            return {}
        payload = json.loads(m.group(1))
        # payload 格式: [{"status":0,"hq":[["2024-01-02","2975.31","2962.85",...], ...]}]
        hq = payload[0].get('hq', [])
        # 字段顺序: [日期, 开盘, 收盘, 涨跌额, 涨跌幅%, 最低, 最高, 成交量, 成交额, 换手率]
        result = {}
        for row in hq:
            if len(row) >= 3:
                date_str = row[0]   # YYYY-MM-DD
                try:
                    close = float(str(row[2]).replace(',', ''))
                    result[date_str] = close
                except (ValueError, TypeError):
                    pass
        print(f'  [上证指数] 获取 {len(result)} 个交易日收盘点位')
        return result
    except Exception as e:
        print(f'  [上证指数] 请求失败: {e}')
        return {}


def load_shindex_checkpoint():
    """读取上证指数本地缓存，返回 {date: close} 字典。"""
    if not os.path.exists(SHINDEX_CHECKPOINT):
        return {}
    try:
        with open(SHINDEX_CHECKPOINT, 'r', encoding='utf-8') as f:
            data = json.load(f)
        prices = data.get('prices', {})
        print(f'  [上证指数] 本地缓存：{len(prices)} 个交易日，最新 {max(prices) if prices else "无"}')
        return prices
    except Exception:
        return {}


def save_shindex_checkpoint(prices):
    """原子写入上证指数缓存。"""
    if not prices:
        return
    tmp = SHINDEX_CHECKPOINT + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump({
            'count':    len(prices),
            'latest':   max(prices),
            'earliest': min(prices),
            'saved_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'prices':   prices,
        }, f, ensure_ascii=False)
    os.replace(tmp, SHINDEX_CHECKPOINT)


def get_shindex_data():
    """
    智能获取上证指数数据：
    - 本地有缓存且最新日期为今天 → 直接用缓存
    - 否则 → 重新从搜狐拉取全量并更新缓存
    返回 {'YYYY-MM-DD': close_price} 字典
    """
    cached = load_shindex_checkpoint()
    today  = today_str()
    latest = max(cached) if cached else None

    # 搜狐接口一次返回全量，直接重拉即可（通常<1秒）
    if latest and latest >= today:
        print(f'  [上证指数] 数据已是最新（{latest}），使用本地缓存。')
        return cached

    print(f'  [上证指数] 从搜狐接口拉取最新数据...')
    fresh = fetch_shindex()
    if fresh:
        # 合并新旧（以防搜狐接口覆盖不完整）
        cached.update(fresh)
        save_shindex_checkpoint(cached)
        return cached
    elif cached:
        print(f'  [上证指数] 接口失败，使用本地缓存（最新 {latest}）。')
        return cached
    else:
        print(f'  [上证指数] 无法获取数据，跳过上证指数。')
        return {}


# ── 8. Excel 导出 ──────────────────────────────────────────────────────────────

HDR_FILL  = PatternFill('solid', start_color='1F4E79')
SUB_FILL  = PatternFill('solid', start_color='2E75B6')
ALT_FILL  = PatternFill('solid', start_color='EBF3FB')
HDR_FONT  = Font(name='Arial', bold=True, color='FFFFFF', size=13)
SUB_FONT  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
BODY_FONT = Font(name='Arial', size=10)
CENTER    = Alignment(horizontal='center', vertical='center')
LEFT_ALGN = Alignment(horizontal='left',   vertical='center')
NUM_FMT   = '#,##0.00'
THIN      = Side(style='thin', color='BDD7EE')
BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def sc(cell, font=None, fill=None, alignment=None, number_format=None, border=None):
    if font:          cell.font          = font
    if fill:          cell.fill          = fill
    if alignment:     cell.alignment     = alignment
    if number_format: cell.number_format = number_format
    if border:        cell.border        = border


def generate_excel(plot_data, output_path, index_prices=None):
    wb        = Workbook()
    all_dates = sorted({d for t in plot_data for d in t['x']})
    etf_names = [t['name'] for t in plot_data]

    # Sheet1 透视表
    ws1 = wb.active
    ws1.title = '透视表（日期×ETF）'
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(etf_names)+1)
    sc(ws1.cell(row=1, column=1, value='上交所宽基 ETF 规模历史数据（单位：万份）'),
       font=HDR_FONT, fill=HDR_FILL, alignment=CENTER)
    ws1.row_dimensions[1].height = 30
    sc(ws1.cell(row=2, column=1, value='统计日期'),
       font=SUB_FONT, fill=SUB_FILL, alignment=CENTER, border=BORDER)
    for ci, name in enumerate(etf_names, start=2):
        sc(ws1.cell(row=2, column=ci, value=name),
           font=SUB_FONT, fill=SUB_FILL, alignment=CENTER, border=BORDER)
    ws1.row_dimensions[2].height = 22

    d2r = {d: r for r, d in enumerate(all_dates, start=3)}
    for date in all_dates:
        r = d2r[date]; fill = ALT_FILL if r % 2 == 0 else None
        sc(ws1.cell(row=r, column=1, value=date),
           font=BODY_FONT, fill=fill, alignment=CENTER, border=BORDER)
    for ci, trace in enumerate(plot_data, start=2):
        for date, val in zip(trace['x'], trace['y']):
            r = d2r[date]; fill = ALT_FILL if r % 2 == 0 else None
            sc(ws1.cell(row=r, column=ci, value=val),
               font=BODY_FONT, fill=fill, alignment=CENTER,
               number_format=NUM_FMT, border=BORDER)
    ws1.column_dimensions['A'].width = 16
    for ci in range(2, len(etf_names)+2):
        ws1.column_dimensions[get_column_letter(ci)].width = 26
    ws1.freeze_panes = 'A3'

    # Sheet2 明细表
    ws2 = wb.create_sheet('明细表')
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    sc(ws2.cell(row=1, column=1, value='上交所宽基 ETF 规模历史明细'),
       font=HDR_FONT, fill=HDR_FILL, alignment=CENTER)
    ws2.row_dimensions[1].height = 30
    for ci, h in enumerate(['ETF名称','代码','统计日期','规模（万份）'], start=1):
        sc(ws2.cell(row=2, column=ci, value=h),
           font=SUB_FONT, fill=SUB_FILL, alignment=CENTER, border=BORDER)
    ws2.row_dimensions[2].height = 22

    row = 3
    for trace in plot_data:
        name = trace['name'].split('(')[0]
        code = trace['name'].split('(')[1].rstrip(')')
        for date, val in zip(trace['x'], trace['y']):
            fill = ALT_FILL if row % 2 == 0 else None
            for ci, v in enumerate([name, code, date, val], start=1):
                sc(ws2.cell(row=row, column=ci, value=v),
                   font=BODY_FONT, fill=fill,
                   alignment=LEFT_ALGN if ci == 1 else CENTER,
                   number_format=NUM_FMT if ci == 4 else None,
                   border=BORDER)
            row += 1
    for ci, w in enumerate([28, 12, 16, 18], start=1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.freeze_panes = 'A3'

    # Sheet3 上证指数
    if index_prices:
        ws3        = wb.create_sheet('上证指数')
        idx_headers = ['统计日期', '收盘点位']
        ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        sc(ws3.cell(row=1, column=1, value='上证指数（000001）历史收盘点位'),
           font=HDR_FONT, fill=HDR_FILL, alignment=CENTER)
        ws3.row_dimensions[1].height = 30
        for ci, h in enumerate(idx_headers, start=1):
            sc(ws3.cell(row=2, column=ci, value=h),
               font=SUB_FONT, fill=SUB_FILL, alignment=CENTER, border=BORDER)
        ws3.row_dimensions[2].height = 22
        for irow, (date, price) in enumerate(
                sorted(index_prices.items(), reverse=True), start=3):
            fill = ALT_FILL if irow % 2 == 0 else None
            sc(ws3.cell(row=irow, column=1, value=date),
               font=BODY_FONT, fill=fill, alignment=CENTER, border=BORDER)
            sc(ws3.cell(row=irow, column=2, value=price),
               font=BODY_FONT, fill=fill, alignment=CENTER,
               number_format='#,##0.00', border=BORDER)
        ws3.column_dimensions['A'].width = 16
        ws3.column_dimensions['B'].width = 18
        ws3.freeze_panes = 'A3'

    wb.save(output_path)
    print(f'✅ Excel 已生成：{output_path}  （共 {row-3} 行 ETF 明细）')


# ── 9. HTML 生成 ───────────────────────────────────────────────────────────────

HTML_TEMPLATE = """<!DOCTYPE html>
<html>
<head>
  <meta charset="utf-8">
  <title>上交所宽基 ETF 规模监控</title>
  <script src="https://cdn.plot.ly/plotly-2.27.0.min.js"></script>
  <style>
    body { background:#f5f7fa; color:#333; font-family:'PingFang SC','Segoe UI',sans-serif; margin:0; padding:20px; }
    h2   { text-align:center; color:#2c3e50; font-weight:500; margin-bottom:20px; }
    .tabs-container { display:flex; flex-wrap:wrap; justify-content:center; gap:10px; width:95%; margin:0 auto 20px; }
    .etf-block {
      padding:10px 18px; background:#fff; border:1px solid #dcdfe6;
      border-radius:8px; cursor:pointer; font-size:14px; color:#606266;
      transition:all .2s; box-shadow:0 2px 4px rgba(0,0,0,.02);
    }
    .etf-block:hover  { border-color:#409EFF; color:#409EFF; }
    .etf-block.active { background:#409EFF; color:#fff; border-color:#409EFF; font-weight:bold; box-shadow:0 4px 8px rgba(64,158,255,.3); }
    .chart-container  { width:95%; height:650px; margin:0 auto; background:#fff; padding:20px; border-radius:12px; box-shadow:0 4px 16px rgba(0,0,0,.05); }
  </style>
</head>
<body>
  <h2>上交所宽基 ETF 规模监控看板（DATA_RANGE / 单位：万份）</h2>
  <div class="tabs-container" id="tabs"></div>
  <div class="chart-container" id="chart"></div>
  <script>
    var rawData = PLOT_DATA_JSON;
    var layout = {
      paper_bgcolor:'#fff', plot_bgcolor:'#fff',
      title:{ text:'总体汇总视图', font:{size:18} },
      xaxis:{ title:'统计日期（YYYY-MM-DD）', type:'category', tickmode:'linear', dtick:30, tickangle:-45, gridcolor:'#f0f0f0' },
      yaxis:{ title:'ETF 规模（万份）', gridcolor:'#f0f0f0', side:'left' },
      yaxis2:{
        title:'上证指数（点）',
        overlaying:'y',
        side:'right',
        showgrid:false,
        tickfont:{ color:'#e74c3c' },
        titlefont:{ color:'#e74c3c' }
      },
      hovermode:'x unified',
      legend:{ orientation:'v', x:1.08, y:1 },
      margin:{ t:60, r:200, b:80, l:70 }
    };
    Plotly.newPlot('chart', rawData, layout);
    var tabsDiv = document.getElementById('tabs');
    var btnAll = document.createElement('div');
    btnAll.className = 'etf-block active';
    btnAll.textContent = '📊 总体汇总';
    btnAll.onclick = function() {
      document.querySelectorAll('.etf-block').forEach(b => b.classList.remove('active'));
      this.classList.add('active');
      Plotly.restyle('chart', {visible:true});
      Plotly.relayout('chart', {title:'总体汇总视图'});
    };
    tabsDiv.appendChild(btnAll);
    rawData.forEach(function(trace, idx) {
      var btn = document.createElement('div');
      btn.className = 'etf-block';
      btn.textContent = trace.name.split('(')[0];
      btn.onclick = function() {
        document.querySelectorAll('.etf-block').forEach(b => b.classList.remove('active'));
        this.classList.add('active');
        var mask = rawData.map(function(_, i){ return i === idx; });
        Plotly.update('chart', {visible:mask}, {title:trace.name+' 规模走势'});
      };
      tabsDiv.appendChild(btn);
    });
  </script>
</body>
</html>"""


def generate_html(plot_data, output_path, completed=True, index_prices=None):
    all_dates  = sorted({d for t in plot_data for d in t['x']})
    date_range = f'{all_dates[0]} ~ {all_dates[-1]}' if all_dates else '2020年至今'
    if not completed:
        date_range += '（下载中）'
    html = (HTML_TEMPLATE
            .replace('PLOT_DATA_JSON', json.dumps(plot_data, ensure_ascii=False))
            .replace('DATA_RANGE', date_range))
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f'✅ HTML 已生成：{output_path}')


# ── 10. 主流程 ─────────────────────────────────────────────────────────────────

def main():
    print('=' * 60)
    print('  上交所宽基 ETF 规模监控')
    print('=' * 60)

    today          = today_str()
    existing_dates = get_existing_dates()
    earliest       = min(existing_dates) if existing_dates else None
    latest         = max(existing_dates) if existing_dates else None
    history_done   = (earliest is not None and earliest <= CUTOFF_DATE.strftime('%Y-%m-%d'))

    print(f'📅 当前日期        : {today}')
    print(f'📂 已有数据范围    : {earliest} ~ {latest}（{len(existing_dates)} 个交易日）' if existing_dates else '📂 暂无本地数据')
    print(f'📚 历史已补全至2020: {"是" if history_done else "否"}')
    print()

    # ── Step 1：增量更新（先把最近几天补齐） ──────────────────────────────────
    # 只要 latest 不是今天，就尝试从 latest 的次日到今天补充新数据
    need_increment = (latest is None or latest < today)
    if need_increment:
        print(f'▶  Step 1：增量更新 — 检查 {latest or "无"} 之后到今天（{today}）的新数据')
        print(f'   注意：若今天返回"数据未发布"属正常现象，交易所数据通常在收盘后更新')
        results, new_count = incremental_update()
        if new_count > 0:
            print(f'   本次新增 {new_count} 个交易日。')
        else:
            print('   今天及近期暂无可用新数据（可能是节假日，或数据尚未发布，下次运行会继续重试）。')
        # 增量完成后刷新 existing_dates / earliest
        existing_dates = get_existing_dates()
        earliest       = min(existing_dates) if existing_dates else None
        history_done   = (earliest is not None and earliest <= CUTOFF_DATE.strftime('%Y-%m-%d'))
    else:
        print(f'▶  Step 1：增量更新 — {today} 数据已存在，跳过。')
        results = read_all_results()

    # ── Step 2：历史补全（把数据向过去补到 2020-01-01） ──────────────────────
    if not history_done:
        print(f'\n▶  Step 2：历史补全 — 继续向过去补全至 2020-01-01（当前最早: {earliest}）')
        results, completed = collect_history()
    else:
        print(f'\n▶  Step 2：历史已补全至 2020-01-01，跳过。')
        completed = True

    if not results:
        print('\n❌ 无数据，请检查网络后重试。')
        return

    print(f'\n共 {len(results)} 个交易日，生成输出文件...\n')

    # 获取上证指数数据（独立接口，不影响 ETF 下载流程）
    print('▶  获取上证指数历史数据...')
    index_prices = get_shindex_data()

    plot_data = build_plot_data(results, index_prices=index_prices)
    generate_html(plot_data, OUTPUT_HTML, completed=completed, index_prices=index_prices)
    generate_excel(plot_data, OUTPUT_EXCEL, index_prices=index_prices)

    if completed:
        webbrowser.open(OUTPUT_HTML)
        print('\n🌐 已在浏览器打开 HTML 看板。')
    else:
        print('\n📊 已生成当前数据预览，下次运行将继续历史补全。')


if __name__ == '__main__':
    main()